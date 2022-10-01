import openpyxl as opx
import openpyxl.styles as opxs
import os

# scan and print directory files
file_list = []
for path, folders, files in os.walk("."):
    for file in files:
        file_list.append(os.path.join(path, file))

for filename in file_list:
    print(filename)

# input
file_txt_name = input("change file txt: ")
file_exl_name = input("change file xls: ")


# open xls file
wb = opx.load_workbook(file_exl_name)
print(wb.sheetnames)
sheet_name = input("change sheet name: ")
ws = wb[sheet_name]


# open txt file
file_txt = open(file_txt_name)
lines = file_txt.readlines()

# main cycle
count = 1
for line in lines:    
    # print(line.strip())
    group = line.split("->")
    group = group[4].strip()
    # print("group = " + group)
    for i in range(1,len(ws['D']),1):
        cell = ws.cell(i, 4)
        # print("cell = " + str(cell.value))
        if cell.value == group:            
            # cell.fill = opxs.PatternFill(fill_type='solid', start_color='ff0000', end_color='ff0000')
            print(str(count) + ". " + group + " -> " + cell.value + " -> delete")
            ws.delete_rows(i, 1)
            count += 1
wb.save(file_exl_name)
print("finish! delete " + str(count - 1) + " rows")